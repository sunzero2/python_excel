import mysql.connector
import openpyxl

def insert_value(table, values) :
  sql = "INSERT INTO "
  if table == "type" :
    sql += "TYPE (NAME) VALUES ('{}')"
  else :
    sql += "HARDWARE (TYPE_ID, BRAND, MODEL, CAPACITY) VALUES ({typeId}, '{brand}', '{model}', {capacity})"

  print(sql.format(typeId=values[0], brand=values[1], model=values[2], capacity=values[3]))
  # mycursor.execute()

mydb = mysql.connector.connect (
  host="127.0.0.1",
  user="root",
  password="1234",
  database="ngle_pc"
)

brands = ["LG", "SAMSUNG", "DELL", "ALPHASCAN", "RETINA", "INTEL", "NVIDIA", "RADEON", "WDC", "SANDISK", "ADATA", "CRUCIAL", "APPLE"]
mycursor = mydb.cursor()
wb = openpyxl.load_workbook('/Users/leehyeyeong/Downloads/nGle_pc_final.xlsx')
sheets = wb.sheetnames

for i in sheets:
  worksheet = wb[i]
  for row in worksheet.iter_rows(max_col=1) :
    if (i == "type") :
      # print(" ")
      insert_value(i, row[0].value)
    else :
      tmp = row[0].value.upper().split(" ")
      brand = tmp[0] if brands.count(tmp[0]) > 0 else "UNKNOWN"
      
      v = [wb.worksheets.index(worksheet), brand, "MODEL", "CAPACITY"]
      insert_value(i, v)
  # mydb.commit()


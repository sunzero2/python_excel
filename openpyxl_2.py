import openpyxl

# 새로운 엑셀 파일 생성
# wb = openpyxl.Workbook()

# 엑셀 파일 가져오기
wb = openpyxl.load_workbook('/Users/leehyeyeong/Downloads/nGle_PC.xlsx')

# 활성화되어 있는 시트 가져옴
sheet2 = wb.worksheets[0]

# 새로운 시트 생성
# sheet2 = wb.create_sheet('통합2')

# 시트 불러오기
# sheet3 = wb['통합']

sheet2.title = '시트 이름 변경'
sheet2.append([1, 2, 3, 4, 5])
# sheet2.cell(row=3, column=3).value = '3, 3'
wb.save('/Users/leehyeyeong/Downloads/nGle_PC.xlsx')